import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, WebDriverException, SessionNotCreatedException, ElementNotInteractableException, StaleElementReferenceException
from selenium.common.exceptions import TimeoutException
import numpy as np
import pandas as pd
from fake_useragent import UserAgent
from datetime import date, datetime
import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

def write_csv(my_dict):

    if os.path.exists("l2l_checklist.xlsx"):
        workbook = load_workbook('l2l_checklist.xlsx')
        sheet = workbook.active
        start_row = sheet.max_row + 1
        column = 1
        for key, value in my_dict.items():
            sheet.cell(row=start_row, column=column).value = value
            column += 1

        for row in sheet.iter_rows():
            for cell in row:
                # Apply 'wrap_text' formatting
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            # Find the maximum length of content in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save('l2l_checklist.xlsx')
        
    else:
        workbook = Workbook()
        sheet = workbook.active
        column = 1
        for key, value in my_dict.items():
            sheet.cell(row=1, column=column).value = key
            sheet.cell(row=2, column=column).value = value
            column += 1

        for row in sheet.iter_rows():
            for cell in row:
                # Apply 'wrap_text' formatting
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            # Find the maximum length of content in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

        # Set the width of the column to fit the maximum length of content
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save('l2l_checklist.xlsx')

def initialize_user_agent_and_ip_rotation():
    useragent_obj = UserAgent(browsers=["edge", "firefox", "safari", 'chrome'], use_external_data=True)
    useragent = str(useragent_obj.random)
    PROXY = "http://api.scraperapi.com?api_key={}&url=http://httpbin.org/ip&render=true".format("b85d057a0618675b026177fb3351ea6d")
    # PROXY = "65.109.160.214:8080"
    return(useragent, PROXY)

def required_options_and_Driver(type_of_driver=True):
    # chromedriver_autoinstaller.install()

    # few error handling
    options = webdriver.ChromeOptions()
    # options.add_argument('--headless=new')
    # options.add_experimental_option("detach", True)
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-browser-side-navigation")
    options.add_argument('--disable-blink-features=AutomationControlled')
    # options.add_argument("--user-data-dir=" + r"C:\Users\artayade\AppData\Local\Google\Chrome\User Data\Default")
    # options.add_argument("--incognito")

    useragent, PROXY = initialize_user_agent_and_ip_rotation()

    # user agent rotation
    options.add_argument("user-agent={}".format(useragent))
    options.add_argument("--start-maximized")
    # ip rotation

    options.add_argument("--proxy-server={}".format(PROXY))

    if type_of_driver == True:
        driver = uc.Chrome(options=options, driver_executable_path='chromedriver/chromedriver.exe')
    else:
        driver = webdriver.Chrome(options=options)

    driver.delete_all_cookies()

    return(driver)

def run(df):

    already_scraped_ids = []

    # if file exists
    if os.path.exists('l2l_checklist.xlsx'):
        df_scraped = pd.read_excel('l2l_checklist.xlsx')
        already_scraped_ids = df_scraped['Checklist_Id'].to_list()
        print(already_scraped_ids)


    driver = required_options_and_Driver()
    
    driver.get("https://wi.leading2lean.com/home/")
    driver.implicitly_wait(5)
    
    sleep(2)

    #enter username
    usr_nm = driver.find_element(By.ID, 'username')
    usr_nm.send_keys('209203')
    sleep(1)

    password = driver.find_element(By.ID, 'password')
    password.send_keys('Worthington$123')
    sleep(1)
    password.send_keys(Keys.ENTER)

    sleep(2)

    previous_value = None

    for i, row in df.iterrows():

        if row['document'] not in already_scraped_ids:

            my_dict = {}

            current_value = row['site']

            if current_value != previous_value:
                previous_value = current_value
                # print(f"{previous_value} changed")
                driver.get(f'https://wi.leading2lean.com/selectsite/{previous_value}/?next=')
                driver.implicitly_wait(2)
            else:
                driver.get(f'https://wi.leading2lean.com/selectsite/{previous_value}/?next=')
                driver.implicitly_wait(2)

            # driver.get(f'https://wi.leading2lean.com/selectsite/6/?next=')
            sleep(2)

            driver.get(f"https://wi.leading2lean.com/documents/checklist_templates/{row['document']}/{row['revision']}/?closetab=False")
            # driver.get(f"https://wi.leading2lean.com/documents/checklist_templates/{20832}/{28332}/?closetab=False")
            driver.implicitly_wait(2)

            sleep(2)

            my_dict['Site'] = row['site']
            my_dict['Checklist_Id'] = row['document']
            my_dict['Revision'] = row['revision']

            content = driver.find_element(By.CSS_SELECTOR, '#checklist_form > table > tbody > tr:nth-child(3)').text.split("\n")

            prod_id = content.index('Production Settings:   ')
            reject_limit_id = content.index('Launch a Dispatch if values exceed the Reject Limits:   ')

            production_settings = content[prod_id:reject_limit_id]
            dispatch_settings = content[reject_limit_id:]

            # selected answer for Work Schedule Type
            work_schedule_type_Selected = production_settings[production_settings.index('Repeat Daily', 1)+1:]
            # print(work_schedule_type_Selected)
            my_dict['Stand_Work_Schedule_Type'] = work_schedule_type_Selected[0]

            if work_schedule_type_Selected[0] == 'Repeat Daily':
                #time of the day
                time_of_day = driver.find_element(By.ID, 'repeat_time_of_day').get_attribute('value')
                my_dict['Interval'] = np.NAN
                my_dict['Minutes_From'] = np.NAN
                my_dict['Time_of_Day'] = time_of_day
            
            elif work_schedule_type_Selected[0] == 'Repeat':
                # repeat every minutes
                interval = driver.find_element(By.ID, 'repeat_every_x_minutes').get_attribute('value')
                my_dict['Interval'] = interval
                my_dict['Minutes_From'] = np.NAN
                my_dict['Time_of_Day'] = np.NAN
            
            elif work_schedule_type_Selected[0] == 'Shift Boundary':
                # repeat every minutes
                interval = driver.find_element(By.ID, 'repeat_every_x_minutes').get_attribute('value')
                my_dict['Interval'] = interval

                Minutes_From = driver.find_element(By.ID, 'id_shift_due_type').get_attribute('value')
                # Minutes_From - returns either 0 or 1

                if Minutes_From == 0:
                    my_dict['Minutes_From'] = 'Beginning of shift'
                else:
                    my_dict['Minutes_From'] = 'End of shift'
                
                my_dict['Time_of_Day'] = np.NAN

                # 0 - Beginning of shift, 1 - End of shift

            elif work_schedule_type_Selected[0] == 'None':
                my_dict['Interval'] = np.NAN
                my_dict['Minutes_From'] = np.NAN
                my_dict['Time_of_Day'] = np.NAN
            
            if 'This checklist is currently shared' in dispatch_settings[0]:
                my_dict['Dispatch_Type'] = np.NAN
                my_dict['Trade'] = np.NAN
                # print("NULL")

            else:
                #dispatch type
                Dispatch_Type = driver.find_elements(By.CLASS_NAME, 'twocol-first-auto')[1].get_attribute('value')
                #trade type
                Trade = driver.find_elements(By.CLASS_NAME, 'twocol-first-auto')[2].get_attribute('value')
                # print(dispatch_settings)

                my_dict['Dispatch_Type'] = Dispatch_Type
                my_dict['Trade'] = Trade

            # print(my_dict)
            write_csv(my_dict)
            sleep(2)
        
        else:
            continue
        
    driver.quit() 

def main():
    df = read_checklist()
    run(df)

def read_checklist():

    df = pd.read_excel(r"Z:\Cylinders Analytics\Checklist Scrapping\checklist list.xlsx")
    # print(df['site'].unique().tolist())

    return(df)

if __name__ == '__main__':

    # main()
    # print(read_checklist())

    df1 = pd.read_excel('l2l_checklist.xlsx')
    print(len(df1['Checklist_Id'].unique()))

    df2 = read_checklist()
    print(len(df2['document'].unique()))

